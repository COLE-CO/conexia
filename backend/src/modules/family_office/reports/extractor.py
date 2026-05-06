import io

import openpyxl
import pdfplumber

MAX_CHARS = 24000


def extract_balance_content(file_bytes: bytes, file_type: str) -> str:
    if file_type in (".xlsx", ".xls"):
        return _extract_xlsx(file_bytes)
    if file_type == ".pdf":
        return _extract_pdf(file_bytes)
    return ""


def _extract_xlsx(file_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"--- Hoja: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append(" | ".join(cells))

    wb.close()
    text = "\n".join(lines)
    return text[:MAX_CHARS]


def _extract_pdf(file_bytes: bytes) -> str:
    lines: list[str] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        cells = [str(c) if c else "" for c in row]
                        lines.append(" | ".join(cells))
            else:
                text = page.extract_text()
                if text:
                    lines.append(text)

    text = "\n".join(lines)
    return text[:MAX_CHARS]
